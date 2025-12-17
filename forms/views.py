from django.shortcuts import render
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework import viewsets
from rest_framework.decorators import action
from datetime import datetime
from io import BytesIO
# Create your views here.
from rest_framework import viewsets, filters
from django_filters.rest_framework import DjangoFilterBackend
from .models import ACFailureReport, FailureReport, IPSReport, JPCReport, MaintenanceReport, MovementReport, RelayRoomLog
from .serializers import ACFailureReportSerializer, FailureReportSerializer, IPSReportSerializer, JPCReportSerializer, MaintenanceReportSerializer, MovementReportSerializer, RelayRoomLogSerializer
from django_filters import rest_framework as django_filters
from rest_framework.filters import SearchFilter, OrderingFilter

class RelayRoomLogFilter(django_filters.FilterSet):
    csi = django_filters.CharFilter(field_name='csi_unit__name', lookup_expr='exact')
    sectional_officer = django_filters.CharFilter(field_name='sectional_officer__name', lookup_expr='exact')
    location__code = django_filters.CharFilter(field_name='location__code', lookup_expr='exact')
    date = django_filters.DateFilter(field_name='log_date')

    class Meta:
        model = RelayRoomLog
        fields = ['location']

class ACFailureReportFilter(django_filters.FilterSet):
    csi = django_filters.CharFilter(field_name='csi_unit__name', lookup_expr='exact')
    sectional_officer = django_filters.CharFilter(field_name='sectional_officer__name', lookup_expr='exact')
    
    class Meta:
        model = ACFailureReport
        fields = ['location_code', 'under_warranty', 'under_amc']

class MaintenanceReportViewSet(viewsets.ModelViewSet):
    queryset = MaintenanceReport.objects.all().order_by('-submission_date')
    serializer_class = MaintenanceReportSerializer
    
    # Optional: If you need to filter by date or officer later
    filterset_fields = ['sectional_officer__name', 'csi_unit__name', 'submission_date']


IPS_MODULES = [
    'SMR', 'Inverter', 'DC-DC Converter', 'Transformer', 
    'AVR/CVT', 'CSU', 'SM Status Monitoring Panel'
]

IPS_COMPANIES = [
    'AMARARAJA', 'HBL', 'STATCON', 'STATCON INERTIA', 'SUKHILA'
]

class IPSReportViewSet(viewsets.ModelViewSet):
    # Prefetch related fields to prevent N+1 SQL query issues
    queryset = IPSReport.objects.all().select_related('csi_unit').prefetch_related(
        'entries__module_type', 
        'entries__company'
    ).order_by('-submission_date')
    serializer_class = IPSReportSerializer
    
    # Filtering capabilities
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['csi_unit__name', 'submission_date'] # Filter by CSI name or date
    ordering_fields = ['submission_date', 'created_at'] # Allow ordering by date fields

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        # 1. Filter Data
        queryset = self.filter_queryset(self.get_queryset())
        
        # 2. Setup Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "IPS Position"

        # --- CONSTANTS ---
        # Ensure these names match your Company Model 'name' fields exactly
        IPS_COMPANIES = ['AMARARAJA', 'HBL', 'STATCON', 'STATCON INERTIA', 'SUKHILA']
        
        # Exact order from the PDF
        IPS_MODULES = [
            'SMR', 
            'Inverter', 
            'DC-DC Converter', 
            'Transformer', 
            'AVR/CVT', 
            'CSU', 
            'SM status monitoring panel'
        ]

        # --- STYLES ---
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        # Fonts
        title_font = Font(bold=True, size=14, name='Calibri')
        header_font = Font(bold=True, size=10, name='Calibri')
        data_font = Font(size=10, name='Calibri')
        bold_data_font = Font(bold=True, size=10, name='Calibri')
        
        # Alignment
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Colors
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Light Grey
        csi_total_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid") # Darker Grey
        grand_total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Light Yellow (optional, or keep white)

        # --- ROW 1: MAIN TITLE ---
        # Calculate total columns: 3 (Sr, CSI, Mod) + (5 Companies * 4 cols) + (1 Total * 4 cols) + 1 (Remarks) = 28 columns
        total_cols = 3 + (len(IPS_COMPANIES) * 4) + 4 + 1
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"IPS modules defective and spare position as on {datetime.now().strftime('%d.%m.%Y')}"
        title_cell.font = title_font
        title_cell.alignment = center_align
        title_cell.border = thin_border

        # --- ROW 2 & 3: HEADERS ---
        # Static Columns
        ws.merge_cells('A2:A3')
        ws['A2'] = 'Sr.No'
        ws.merge_cells('B2:B3')
        ws['B2'] = 'CSI'
        ws.merge_cells('C2:C3')
        ws['C2'] = 'Details of faulty modules'

        current_col = 4
        # Company Headers
        for company in IPS_COMPANIES:
            end_col = current_col + 3
            cell_ref = f"{get_column_letter(current_col)}2:{get_column_letter(end_col)}2"
            ws.merge_cells(cell_ref)
            cell = ws.cell(row=2, column=current_col)
            cell.value = company.upper()
            cell.font = header_font
            cell.alignment = center_align
            cell.fill = header_fill
            current_col = end_col + 1

        # Total Header
        end_col = current_col + 3
        cell_ref = f"{get_column_letter(current_col)}2:{get_column_letter(end_col)}2"
        ws.merge_cells(cell_ref)
        cell = ws.cell(row=2, column=current_col)
        cell.value = "TOTAL"
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        
        # Remarks Header
        remarks_col = end_col + 1
        ws.merge_cells(f"{get_column_letter(remarks_col)}2:{get_column_letter(remarks_col)}3")
        cell = ws.cell(row=2, column=remarks_col)
        cell.value = "Action Plan / Remarks"
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

        # Row 3: Sub-headers
        # Per PDF: Def, Spare, Spare mod under AMC/ARC, Def mod under AMC/ARC
        sub_headers = ['Def.', 'Spare', 'Spare\nmod\nunder\nAMC/\nARC', 'Def mod\nunder\nAMC/\nARC']
        
        col_idx = 4
        for _ in range(len(IPS_COMPANIES) + 1): # Companies + Total Column
            for header in sub_headers:
                cell = ws.cell(row=3, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, size=8, name='Calibri')
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = header_fill
                col_idx += 1

        # Apply borders/styles to Row 2 static columns
        for row in ws.iter_rows(min_row=2, max_row=3, max_col=remarks_col):
            for cell in row:
                cell.border = thin_border
                if cell.row == 2 and cell.column <= 3: # SrNo, CSI, Details Headers
                    cell.alignment = center_align
                    cell.fill = header_fill
                    cell.font = header_font

        # --- DATA INITIALIZATION ---
        row_num = 4
        sr_no = 1
        
        # Dictionary to store Grand Totals (Vertical sum of everything)
        # Structure: {'AMARARAJA': [def, spare, def_amc, spare_amc], ... 'TOTAL': [...]}
        grand_totals = {comp: [0, 0, 0, 0] for comp in IPS_COMPANIES}
        grand_totals['TOTAL'] = [0, 0, 0, 0]

        # Dictionary to store Module-wise Grand Totals
        # Structure: {'SMR': {'AMARARAJA': [0,0,0,0], ... 'TOTAL': [0,0,0,0]}, ...}
        global_module_totals = {
            mod: {comp: [0, 0, 0, 0] for comp in IPS_COMPANIES + ['TOTAL']} 
            for mod in IPS_MODULES
        }

        for report in queryset:
            start_row = row_num
            report_entries = list(report.entries.all())

            # Local totals for this specific CSI (to be written in the grey row)
            # Structure: {'AMARARAJA': [def, spare, def_amc, spare_amc], ...}
            csi_column_totals = {comp: [0, 0, 0, 0] for comp in IPS_COMPANIES}
            csi_row_total_accumulator = [0, 0, 0, 0] # For the horizontal Total column of the CSI total row

            # 1. Iterate through specific modules
            for module_name in IPS_MODULES:
                # Column 3: Module Name
                ws.cell(row=row_num, column=3).value = module_name
                ws.cell(row=row_num, column=3).font = data_font
                ws.cell(row=row_num, column=3).border = thin_border

                col_idx = 4
                
                # Accumulator for this specific row (Horizontal sum across companies)
                row_horizontal_sum = [0, 0, 0, 0] # [def, spare, def_amc, spare_amc]

                for company in IPS_COMPANIES:
                    # Find entry
                    entry = next((e for e in report_entries if e.module_type.name == module_name and e.company.name == company), None)
                    
                    # Get values
                    val_def = entry.qty_defective if entry else 0
                    val_spare = entry.qty_spare if entry else 0
                    val_def_amc = entry.qty_defective_amc if entry else 0 
                    val_spare_amc = entry.qty_spare_amc if entry else 0

                    # Swapped order: Def, Spare, Spare AMC, Def AMC
                    vals = [val_def, val_spare, val_spare_amc, val_def_amc]

                    # Write 4 cells for this company
                    for i, val in enumerate(vals):
                        c = ws.cell(row=row_num, column=col_idx + i)
                        c.value = val
                        c.alignment = center_align
                        c.border = thin_border
                        c.font = data_font
                        
                        # Add to Row Horizontal Sum
                        row_horizontal_sum[i] += val
                        # Add to CSI Column Total
                        csi_column_totals[company][i] += val
                        # Add to Grand Total
                        grand_totals[company][i] += val
                        # Add to Global Module Total
                        global_module_totals[module_name][company][i] += val

                    col_idx += 4

                # Write Horizontal Row Total (The "TOTAL" column group at right)
                for i, val in enumerate(row_horizontal_sum):
                    c = ws.cell(row=row_num, column=col_idx + i)
                    c.value = val
                    c.alignment = center_align
                    c.border = thin_border
                    c.font = bold_data_font # Make row totals bold
                    c.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid") # Light separation
                    
                    # Accumulate for the CSI Total Row
                    csi_row_total_accumulator[i] += val
                    # Accumulate for Grand Total of Totals
                    grand_totals['TOTAL'][i] += val
                    # Accumulate for Global Module Total
                    global_module_totals[module_name]['TOTAL'][i] += val

                row_num += 1

            # 2. CSI TOTAL ROW (The Grey Row)
            ws.cell(row=row_num, column=3).value = "TOTAL"
            ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=row_num, column=3).font = bold_data_font
            ws.cell(row=row_num, column=3).fill = csi_total_fill
            ws.cell(row=row_num, column=3).border = thin_border

            col_idx = 4
            for company in IPS_COMPANIES:
                vals = csi_column_totals[company]
                for i, val in enumerate(vals):
                    c = ws.cell(row=row_num, column=col_idx + i)
                    c.value = val
                    c.font = bold_data_font
                    c.alignment = center_align
                    c.border = thin_border
                    c.fill = csi_total_fill
                col_idx += 4
            
            # Write CSI Total's Total
            for i, val in enumerate(csi_row_total_accumulator):
                c = ws.cell(row=row_num, column=col_idx + i)
                c.value = val
                c.font = bold_data_font
                c.alignment = center_align
                c.border = thin_border
                c.fill = csi_total_fill

            # Style Remarks cell for Total Row
            ws.cell(row=row_num, column=remarks_col).border = thin_border
            ws.cell(row=row_num, column=remarks_col).fill = csi_total_fill

            # 3. MERGING (Sr No, CSI, Remarks)
            # Merge Sr No (spans modules + total row? No, usually spans just modules, but let's look at PDF)
            # In PDF, "TOTAL" row is separate. "Sr No" and "CSI" merges stop before "TOTAL" row.
            
            # Merge Sr No
            ws.merge_cells(start_row=start_row, start_column=1, end_row=row_num-1, end_column=1)
            c = ws.cell(row=start_row, column=1)
            c.value = sr_no
            c.alignment = center_align
            c.font = bold_data_font
            
            # Merge CSI Name
            ws.merge_cells(start_row=start_row, start_column=2, end_row=row_num-1, end_column=2)
            c = ws.cell(row=start_row, column=2)
            c.value = report.csi_unit.name if report.csi_unit else '-'
            c.alignment = center_align
            c.font = bold_data_font

            # Apply borders to merged area
            for r in range(start_row, row_num):
                ws.cell(row=r, column=1).border = thin_border
                ws.cell(row=r, column=2).border = thin_border

            # Remarks (Merge across all rows including Total? Usually Remarks are per CSI)
            ws.merge_cells(start_row=start_row, start_column=remarks_col, end_row=row_num-1, end_column=remarks_col)
            c = ws.cell(row=start_row, column=remarks_col)
            c.value = report.remarks
            c.alignment = left_align
            c.font = Font(italic=True, size=9)
            # Border for remarks
            for r in range(start_row, row_num):
                ws.cell(row=r, column=remarks_col).border = thin_border

            # Handle the empty cells in SrNo/CSI for the Total Row
            ws.cell(row=row_num, column=1).border = thin_border
            ws.cell(row=row_num, column=1).fill = csi_total_fill
            ws.cell(row=row_num, column=2).border = thin_border
            ws.cell(row=row_num, column=2).fill = csi_total_fill

            row_num += 1
            sr_no += 1

        # --- GRAND TOTAL BLOCK ---
        start_grand_total_row = row_num
        
        # 1. Iterate through modules for Grand Total Block
        for module_name in IPS_MODULES:
            # Column 3: Module Name
            ws.cell(row=row_num, column=3).value = module_name
            ws.cell(row=row_num, column=3).font = data_font
            ws.cell(row=row_num, column=3).border = thin_border
            
            col_idx = 4
            for company in IPS_COMPANIES:
                vals = global_module_totals[module_name][company]
                for i, val in enumerate(vals):
                    c = ws.cell(row=row_num, column=col_idx + i)
                    c.value = val
                    c.alignment = center_align
                    c.border = thin_border
                    c.font = data_font
                col_idx += 4
                
            # Horizontal Total for this module row
            vals = global_module_totals[module_name]['TOTAL']
            for i, val in enumerate(vals):
                c = ws.cell(row=row_num, column=col_idx + i)
                c.value = val
                c.alignment = center_align
                c.border = thin_border
                c.font = bold_data_font
                c.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
                
            row_num += 1

        # 2. GRAND TOTAL ROW (Sum of all modules)
        ws.cell(row=row_num, column=3).value = "GRAND TOTAL"
        ws.cell(row=row_num, column=3).font = bold_data_font
        ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=row_num, column=3).fill = grand_total_fill
        ws.cell(row=row_num, column=3).border = thin_border

        col_idx = 4
        # Write Company Grand Totals
        for company in IPS_COMPANIES:
            vals = grand_totals[company]
            for i, val in enumerate(vals):
                c = ws.cell(row=row_num, column=col_idx + i)
                c.value = val
                c.font = bold_data_font
                c.alignment = center_align
                c.border = thin_border
                c.fill = grand_total_fill # Highlight Grand Total
            col_idx += 4
        
        # Write Final Grand Total
        vals = grand_totals['TOTAL']
        for i, val in enumerate(vals):
            c = ws.cell(row=row_num, column=col_idx + i)
            c.value = val
            c.font = bold_data_font
            c.alignment = center_align
            c.border = thin_border
            c.fill = grand_total_fill # Slightly different color for the final block
        
        # Merge Sr No and CSI for Grand Total Block
        ws.merge_cells(start_row=start_grand_total_row, start_column=1, end_row=row_num, end_column=2)
        c = ws.cell(row=start_grand_total_row, column=1)
        c.value = "GRAND TOTAL"
        c.font = bold_data_font
        c.alignment = center_align
        c.fill = grand_total_fill
        
        # Apply borders to merged area
        for r in range(start_grand_total_row, row_num + 1):
            ws.cell(row=r, column=1).border = thin_border
            ws.cell(row=r, column=2).border = thin_border
        
        # Empty border for remarks on last row
        ws.cell(row=row_num, column=remarks_col).border = thin_border

        # --- FINAL ADJUSTMENTS ---
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions[get_column_letter(remarks_col)].width = 30
        
        # Set widths for data columns to be compact
        for i in range(4, remarks_col):
            ws.column_dimensions[get_column_letter(i)].width = 6

        # Output
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"IPS_Position_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    
class RelayRoomLogViewSet(viewsets.ModelViewSet):
    queryset = RelayRoomLog.objects.all().order_by('-log_date')
    serializer_class = RelayRoomLogSerializer
    
    # Enable Filtering
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = RelayRoomLogFilter
    ordering_fields = ['log_date', 'created_at']

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        # 1. Apply Filters (Same as List View)
        queryset = self.filter_queryset(self.get_queryset())
        
        # 2. Setup Excel Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relay Room Data"
        
        # 3. Styles (Minimal/Normal)
        # No background colors, black text
        header_font = Font(name='Calibri', size=11, bold=False) 
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 4. Title Row (Row 1)
        # Construct title based on filters
        date_gte = request.query_params.get('log_date__gte')
        date_lte = request.query_params.get('log_date__lte')
        
        title_text = "Relay room opening position"
        if date_gte and date_lte:
            if date_gte == date_lte:
                title_text += f" on dt:-{datetime.strptime(date_gte, '%Y-%m-%d').strftime('%d/%m/%Y')}"
            else:
                title_text += f" from {datetime.strptime(date_gte, '%Y-%m-%d').strftime('%d/%m/%Y')} to {datetime.strptime(date_lte, '%Y-%m-%d').strftime('%d/%m/%Y')}"
        elif date_gte:
             title_text += f" from {datetime.strptime(date_gte, '%Y-%m-%d').strftime('%d/%m/%Y')}"
        elif date_lte:
             title_text += f" upto {datetime.strptime(date_lte, '%Y-%m-%d').strftime('%d/%m/%Y')}"
        else:
             title_text += f" (All Dates)"

        ws.merge_cells('A1:N1')
        cell = ws.cell(row=1, column=1, value=title_text)
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # No border for title row in the example, but let's keep it clean
        
        # 5. Headers (Row 2)
        headers = [
            'Sr No', 
            'Date', 
            'Station', 
            'Opening time', 
            'Closing time', 
            'Duration in Hours', 
            'Duration in Minutes', 
            'Reason for opening', 
            'CSI', 
            'Sr.No. For open', 
            'Sr. No. For Close', 
            'As per D/L Opening time', 
            'As per D/L Closing time', 
            'CODES for reasons'
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # 6. Data Rows (Start from Row 3)
        for row_idx, log in enumerate(queryset, 3): 
            # Calculate Duration
            duration_str = "-"
            duration_min = 0
            
            if log.opening_time and log.closing_time:
                # Convert time to minutes from midnight
                start_min = log.opening_time.hour * 60 + log.opening_time.minute
                end_min = log.closing_time.hour * 60 + log.closing_time.minute
                
                diff = end_min - start_min
                if diff < 0:
                    diff += 24 * 60 # Handle overnight cases
                
                duration_min = diff
                hours = diff // 60
                mins = diff % 60
                duration_str = f"{str(hours).zfill(2)}:{str(mins).zfill(2)}:00" # Format HH:MM:SS

            row_data = [
                row_idx - 2,                # Sr No (1-based index)
                log.log_date.strftime('%d/%m/%Y') if log.log_date else "", # Date DD/MM/YYYY
                str(log.location.code),          # Station
                log.opening_time.strftime('%H:%M') if log.opening_time else "", # Opening Time HH:MM
                log.closing_time.strftime('%H:%M') if log.closing_time else "", # Closing Time HH:MM
                duration_str,               # Duration in Hours
                duration_min,               # Duration in Minutes
                log.remarks,                # Reason
                log.csi_unit.name if log.csi_unit else "", # CSI
                log.sn_opening,             # Sr.No Open
                log.sn_closing,             # Sr.No Close
                "",                         # As per D/L Opening time (Empty)
                "",                         # As per D/L Closing time (Empty)
                log.opening_code            # CODES for reasons
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # Specific Alignments
                if col_idx == 8: # Reason column
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

        # 7. Adjust Column Widths
        widths = [6, 12, 10, 12, 12, 15, 15, 40, 10, 12, 12, 15, 15, 15]
        for i, w in enumerate(widths, 1):
            col_letter = chr(64 + i) # A, B, C...
            ws.column_dimensions[col_letter].width = w

        # 8. Return Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Relay_Room_Position_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
class ACFailureReportViewSet(viewsets.ModelViewSet):
    queryset = ACFailureReport.objects.all().order_by('created_at')
    serializer_class = ACFailureReportSerializer
    filterset_class = ACFailureReportFilter
    
class FailureReportViewSet(viewsets.ModelViewSet):
    queryset = FailureReport.objects.all().order_by('failure_datetime', 'created_at')
    serializer_class = FailureReportSerializer

    def perform_create(self, serializer):
        # Auto-calculate status or other fields if needed
        serializer.save(status='Open')

class MovementReportViewSet(viewsets.ModelViewSet):
    """
    API endpoint for Daily SI/CSI Movement Reports.
    Endpoint: /api/forms/movement-reports/
    """
    queryset = MovementReport.objects.all()
    serializer_class = MovementReportSerializer
    
    # Enable filtering and searching
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    
    # Allow filtering by specific fields (matches the Dashboard filters)
    filterset_fields = {
        'date': ['exact', 'gte', 'lte'],  # Date ranges
        'sectional_officer': ['exact'],
        'csi': ['exact'],
    }
    
    # Allow searching by text fields
    search_fields = ['name', 'move_to', 'work_done']
    
    ordering_fields = ['date', 'submitted_at']

def export_movement_excel(request):
    # 1. Filter Data based on Request Parameters
    queryset = MovementReport.objects.all().order_by('date', 'sectional_officer', 'csi')

    # Apply Filters from Dashboard
    start_date = request.GET.get('date__gte')
    end_date = request.GET.get('date__lte')
    csi_filter = request.GET.get('csi')
    officer_filter = request.GET.get('sectional_officer')

    if start_date:
        queryset = queryset.filter(date__gte=start_date)
    if end_date:
        queryset = queryset.filter(date__lte=end_date)
    if csi_filter:
        queryset = queryset.filter(csi=csi_filter)
    if officer_filter:
        queryset = queryset.filter(sectional_officer=officer_filter)

    # 2. Setup Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Movement"

    # Styling Constants
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid") # Light Orange from image

    # 3. Create Main Header Row
    # Merging cells A1 to F1 for the main title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    
    # Dynamic Title Date
    date_str = datetime.now().strftime("%d.%m.%Y")
    if start_date:
        date_obj = datetime.strptime(start_date, '%Y-%m-%d')
        date_str = date_obj.strftime("%d.%m.%Y")
    
    title_cell.value = f"Daily movement of SSE & JE Signal of ADI-Division as on Date :{date_str}"
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = center_align
    title_cell.fill = header_fill
    title_cell.border = thin_border

    # 4. Create Column Headers (Row 2 & 3)
    # Row 2
    ws.merge_cells('A2:A3') # Sr No
    ws['A2'] = "Sr. No."
    
    ws.merge_cells('B2:B3') # Name
    ws['B2'] = "Name"
    
    ws.merge_cells('C2:C3') # Designation
    ws['C2'] = "Designation"
    
    ws.merge_cells('D2:E2') # Movement (Spans 2 columns)
    ws['D2'] = "Movement"
    
    ws.merge_cells('F2:F3') # Work Done
    ws['F2'] = "Work Done"

    # Row 3 (Sub-headers for Movement)
    ws['D3'] = "From"
    ws['E3'] = "To"

    # Apply Styling to Header Rows (A2 to F3)
    for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=6):
        for cell in row:
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            # Only apply fill if it's not the blank merged cells
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # 5. Populate Data
    row_num = 4
    for idx, report in enumerate(queryset, 1):
        ws.cell(row=row_num, column=1, value=idx).alignment = center_align
        ws.cell(row=row_num, column=2, value=report.name).alignment = left_align
        ws.cell(row=row_num, column=3, value=report.designation).alignment = center_align
        ws.cell(row=row_num, column=4, value=report.move_from).alignment = center_align
        ws.cell(row=row_num, column=5, value=report.move_to).alignment = center_align
        ws.cell(row=row_num, column=6, value=report.work_done).alignment = left_align

        # Apply borders to data cells
        for col in range(1, 7):
            ws.cell(row=row_num, column=col).border = thin_border
        
        row_num += 1

    # 6. Adjust Column Widths
    ws.column_dimensions['A'].width = 8   # Sr No
    ws.column_dimensions['B'].width = 25  # Name
    ws.column_dimensions['C'].width = 20  # Designation
    ws.column_dimensions['D'].width = 15  # From
    ws.column_dimensions['E'].width = 15  # To
    ws.column_dimensions['F'].width = 60  # Work Done

    # 7. Generate Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Movement_Report_{date_str}.xlsx"'
    wb.save(response)
    return response

class JPCReportViewSet(viewsets.ModelViewSet):
    """
    API endpoint for JPC Done Reports.
    Endpoint: /api/forms/jpc-reports/
    """
    queryset = JPCReport.objects.all()
    serializer_class = JPCReportSerializer
    
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    
    # Filtering options matching the Dashboard
    filterset_fields = {
        'station': ['exact'],
        'jpc_date': ['exact', 'gte', 'lte'],
        'inspection_by': ['exact'],
    }
    
    search_fields = ['station', 'inspector_name']
    ordering_fields = ['jpc_date', 'submitted_at']